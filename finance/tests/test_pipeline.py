"""End to end, offline: folder source -> fake extractor -> ledger -> filing -> forecast -> workbook -> read edits back."""
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from ramfin import db, pipeline
from ramfin.extract.extractor import FakeExtractor
from ramfin.filer import LocalFiler
from ramfin.reports.import_decisions import import_workbook
from ramfin.sources.folder import scan_local_folder
from tests.fixtures import extractions as fx


def test_end_to_end(conn, settings, tmp_path):
    drop = tmp_path / "drop"
    drop.mkdir()
    files = {"IMG_1735.jpeg": fx.RECEIPT_HOME_HARDWARE, "petro.jpg": fx.RECEIPT_NO_JOB, "brandt_1779280.pdf": fx.INVOICE_BRANDT,
             "curbing.pdf": fx.INVOICE_CURBING, "2026-08-01_ED_SMITH_Employee_Timesheet.pdf": fx.TIMESHEET_ED, "advice.pdf": fx.CUSTOMER_PAYMENT_IDL,
             "blurry.jpg": fx.ILLEGIBLE, "td_july.pdf": fx.BANK_STATEMENT_TD}
    for i, name in enumerate(files):
        (drop / name).write_bytes(f"fake-{i}".encode())
    db.insert(conn, "ar_invoices", dict(customer="IDL", invoice_no="600", amount=15120.0, status="Open", expected_date="2026-08-21", created_at=db.now_iso()))
    db.insert(conn, "bank_balances", dict(as_of="2026-08-31", account_key="rbc_chq", balance=17782.0, source="manual", created_at=db.now_iso()))
    db.insert(conn, "bank_balances", dict(as_of="2026-08-31", account_key="td_loc", balance=57812.0, source="manual", created_at=db.now_iso()))

    stats = scan_local_folder(conn, settings, drop, source="camscanner")
    assert stats["new"] == 8
    assert scan_local_folder(conn, settings, drop, source="camscanner")["new"] == 0   # idempotent

    extractor = FakeExtractor(files)
    filer = LocalFiler(tmp_path / "sharepoint")
    ps = pipeline.process_new_documents(conn, settings, extractor, filer)
    assert ps["processed"] == 8 and ps["errors"] == 0 and ps["needs_review"] == 2
    filed = sorted(p.relative_to(tmp_path / "sharepoint").as_posix() for p in (tmp_path / "sharepoint").rglob("*") if p.is_file())
    assert any(p.endswith("06_RECEIPTS/2026/2026-08/260818_HOMEHARDWAREQUESNEL_94-66_240617.jpeg") for p in filed)
    assert any("05_AP_INVOICES/2026-08/260812_BRANDTTRACTOR_INV1779280.pdf" in p for p in filed)
    assert any("01_TIMESHEETS/2026/PP_2026-08-01/2026-08-01_EDSMITH_Timesheet.pdf" in p for p in filed)
    assert any("00_UNFILED_NEEDS_REVIEW" in p for p in filed)
    assert any(p.startswith("projects/01_ACTIVE_PROJECTS/241115_IDL_CONSULTING_2024-2025/") and p.endswith("260821_IDL_EFT_15120-00.pdf") for p in filed)
    assert conn.execute("SELECT COUNT(*) n FROM documents WHERE status='new'").fetchone()["n"] == 0

    rec = pipeline.reconcile(conn, settings)
    assert rec["matching"]["receipt"] == 1     # Home Hardware receipt vs TD statement line

    out = pipeline.weekly(conn, settings, tmp_path / "reports", as_of=date(2026, 8, 31))
    assert Path(out["xlsx"]).exists() and Path(out["markdown"]).exists()
    md = Path(out["markdown"]).read_text()
    assert "13-week view" in md and "Job cost" in md and "240617" in md
    assert "RAM finance" in out["digest_subject"]
    wb = load_workbook(out["xlsx"])
    assert {"DASHBOARD", "FORECAST_13WK", "AP_TRACKER", "AR_TRACKER", "RECEIPTS", "MISSING_RECEIPTS", "JOB_COST", "TIMESHEETS", "ACTIONS", "README"} <= set(wb.sheetnames)

    # Rodney edits the workbook: codes the Petro-Canada receipt, marks Brandt paid, resolves an action
    ws = wb["RECEIPTS"]
    for row in ws.iter_rows(min_row=2):
        if row[2].value and "Petro" in str(row[2].value):
            row[6].value, row[7].value = "241115", "02-300"
    ws = wb["AP_TRACKER"]
    for row in ws.iter_rows(min_row=2):
        if row[1].value and "Brandt" in str(row[1].value):
            row[7].value, row[9].value, row[10].value = "Paid", "2026-09-05", "EFT"
    ws = wb["ACTIONS"]
    first = ws.cell(row=2, column=1).value
    ws.cell(row=2, column=6).value = "resolved"
    ws.cell(row=2, column=7).value = "done"
    edited = tmp_path / "edited.xlsx"
    wb.save(edited)
    st = import_workbook(conn, edited)
    assert st["receipts"] == 1 and st["ap"] == 1 and st["actions"] >= 1
    r = conn.execute("SELECT job_no, cost_code FROM receipts r JOIN vendors v ON v.id=r.vendor_id WHERE v.name LIKE 'Petro%'").fetchone()
    assert (r["job_no"], r["cost_code"]) == ("241115", "02-300")
    assert conn.execute("SELECT status, paid_date FROM ap_invoices WHERE invoice_no='1779280'").fetchone()["status"] == "Paid"
    assert conn.execute("SELECT status FROM action_items WHERE id=?", (first,)).fetchone()["status"] == "resolved"
