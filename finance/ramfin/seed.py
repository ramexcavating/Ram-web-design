"""One-time (re-runnable) seeding from the workbooks RAM already keeps on SharePoint, plus config.

  cost codes  <- 2025_Standard Cost Codes List Rev 1.xlsx (PROJECTS library)
  AP register <- RAM_AP_REGISTER.xlsx, REGISTER and RECEIPTS tabs (finance library)
  AR + bank   <- RAM_CASHFLOW_TOOL.xlsx, AR_TRACKER and WEEKLY_INPUT tabs (finance library)
  vendors     <- config/vendors.csv (committed)
  jobs, debts, recurring, employees, equipment <- config.yaml

Everything is upsert-style: run it twice and nothing doubles. Rows Rodney has since edited in ramfin are left alone
(the register row is only used to CREATE an invoice that does not exist yet).
"""
from __future__ import annotations

import io
import logging
import re
import sqlite3
from datetime import date, datetime

from openpyxl import load_workbook

from . import db
from .rules import vendors as vend
from .rules.filing import parse_date

log = logging.getLogger(__name__)

SECTION_RX = re.compile(r"^([A-Z]{2}|\d{1,2})-000$")
TERMS_RX = re.compile(r"net\s*(\d+)", re.I)


def _num(v) -> float | None:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        f = float(s)
    except ValueError:
        return None
    return -f if neg else f


def _d(v) -> str | None:
    if v in (None, "", "-"):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and 40000 < float(v) < 60000:      # Excel serial
        from datetime import timedelta
        return (date(1899, 12, 30) + timedelta(days=int(v))).isoformat()
    p = parse_date(str(v))
    return p.isoformat() if p else None


# ---------------------------------------------------------------- cost codes
def load_cost_codes_xlsx(conn: sqlite3.Connection, data: bytes) -> int:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    n = 0
    section = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, ""):
            continue
        code = str(row[0]).strip()
        desc = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else ""
        if not re.match(r"^[A-Z0-9]{1,2}-[A-Z0-9]{3}$", code):
            continue
        if SECTION_RX.match(code):
            section = desc.title()
        conn.execute("INSERT INTO cost_codes(code, description, category) VALUES(?,?,?) ON CONFLICT(code) DO UPDATE SET description=excluded.description, category=COALESCE(excluded.category, cost_codes.category)",
                     (code, desc, section))
        n += 1
    conn.commit()
    return n


# ---------------------------------------------------------------- AP register
def _terms_days(t: str | None, default: int = 30) -> int:
    if not t:
        return default
    m = TERMS_RX.search(t)
    if m:
        return int(m.group(1))
    if "receipt" in t.lower():
        return 0
    return default


def import_ap_register(conn: sqlite3.Connection, data: bytes) -> dict[str, int]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    stats = {"ap_new": 0, "ap_seen": 0, "receipts_new": 0, "receipts_seen": 0}
    if "REGISTER" in wb.sheetnames:
        ws = wb["REGISTER"]
        rows = ws.iter_rows(values_only=True)
        hdr = None
        for r in rows:
            if r and r[0] == "ID":
                hdr = [str(c).strip() if c else "" for c in r]
                break
        for r in rows:
            if not r or not r[0] or not str(r[0]).startswith("AP-"):
                continue
            d = dict(zip(hdr, r))
            vendor = (d.get("Vendor") or "").strip()
            if not vendor:
                continue
            doc_type = (d.get("Doc Type") or "Invoice").strip()
            status = (d.get("Payment Status") or "Unpaid").strip()
            if doc_type == "Statement":
                continue
            vid = vend.find_or_create(conn, vendor)
            terms = d.get("Terms")
            conn.execute("UPDATE vendors SET default_terms_days=? WHERE id=? AND default_terms_days=30", (_terms_days(terms), vid))
            inv_no = str(d.get("Invoice / Ref #") or d["ID"]).strip()
            if conn.execute("SELECT 1 FROM ap_invoices WHERE vendor_id=? AND invoice_no=?", (vid, inv_no)).fetchone():
                stats["ap_seen"] += 1
                continue
            cc = d.get("Cost Code")
            code = str(cc).split(" - ")[0].strip() if cc else None
            job = str(d.get("Job #") or "").strip() or None
            if job and not conn.execute("SELECT 1 FROM jobs WHERE job_no=?", (job,)).fetchone():
                job = None
            amt = _num(d.get("Amount (CAD, incl GST)"))
            db.insert(conn, "ap_invoices", dict(
                vendor_id=vid, invoice_no=inv_no, invoice_date=_d(d.get("Invoice Date")), due_date=_d(d.get("Due Date")), amount=amt, gst=None,
                amount_confirmed=1 if str(d.get("Amt Confirmed?") or "").strip().lower() in ("yes", "y") else 0,
                status=status if status in ("Unpaid", "Scheduled", "Partially Paid", "Paid", "Deferred (no-breach)", "Disputed", "On Hold", "Void-Credit", "Reference only") else "Unpaid",
                planned_pay_date=_d(d.get("Planned Pay Date")), paid_date=_d(d.get("Date Paid")), method=(d.get("Method") or None), job_no=job, cost_code=code,
                category="supplier", notes=f"seeded from AP register {d['ID']}. {d.get('Notes') or ''}"[:500], created_at=db.now_iso()))
            stats["ap_new"] += 1
    if "RECEIPTS" in wb.sheetnames:
        ws = wb["RECEIPTS"]
        rows = ws.iter_rows(values_only=True)
        hdr = None
        for r in rows:
            if r and any(("vendor" in str(c).lower() or "merchant" in str(c).lower()) for c in r if c) and any("amount" in str(c).lower() or "total" in str(c).lower() for c in r if c):
                hdr = [str(c).strip() if c else "" for c in r]
                break
        if hdr:
            vcol = next((h for h in hdr if "vendor" in h.lower() or "merchant" in h.lower()), None)
            dcol = next((h for h in hdr if "date" in h.lower()), None)
            acol = next((h for h in hdr if h.lower().startswith("amount") or h.lower() == "total"), None)
            jcol = next((h for h in hdr if h.lower().startswith("job")), None)
            ccol = next((h for h in hdr if h.lower().startswith("cost code")), None)
            fcol = next((h for h in hdr if h.lower() in ("filename", "file")), None)
            for r in rows:
                d = dict(zip(hdr, r))
                if not d.get(vcol):
                    continue
                amt = _num(d.get(acol)) if acol else None
                rdate = _d(d.get(dcol)) if dcol else None
                vid = vend.find_or_create(conn, str(d[vcol]))
                if conn.execute("SELECT 1 FROM receipts WHERE vendor_id=? AND receipt_date IS ? AND ABS(COALESCE(amount,0) - ?) < 0.005", (vid, rdate, amt or 0)).fetchone():
                    stats["receipts_seen"] += 1
                    continue
                job = str(d.get(jcol) or "").strip() or None if jcol else None
                code = str(d.get(ccol) or "").split(" - ")[0].strip() or None if ccol else None
                db.insert(conn, "receipts", dict(vendor_id=vid, receipt_date=rdate, amount=amt, job_no=job if job and conn.execute("SELECT 1 FROM jobs WHERE job_no=?", (job,)).fetchone() else None,
                                                 cost_code=code, description=f"seeded from AP register RECEIPTS tab; file {d.get(fcol) if fcol else ''}"[:500], created_at=db.now_iso()))
                stats["receipts_new"] += 1
    conn.commit()
    return stats


# ---------------------------------------------------------------- cash flow tool
def import_cashflow_tool(conn: sqlite3.Connection, data: bytes, settings) -> dict[str, int]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    stats = {"ar_new": 0, "ar_seen": 0, "balances": 0}
    if "AR_TRACKER" in wb.sheetnames:
        rows = wb["AR_TRACKER"].iter_rows(values_only=True)
        hdr = [str(c).strip() if c else "" for c in next(rows)]
        for r in rows:
            d = dict(zip(hdr, r))
            cust = str(d.get("Customer") or "").strip()
            amt = _num(d.get("Amount"))
            status = str(d.get("Status") or "").strip()
            if not cust or amt is None or cust.upper().startswith("TOTAL") or cust == "NOTE" or status == "Info":
                continue
            inv = str(d.get("Invoice #") or "").strip() or f"{cust[:10]}-{_d(d.get('Invoice date')) or ''}"
            if conn.execute("SELECT 1 FROM ar_invoices WHERE customer=? AND invoice_no=?", (cust, inv)).fetchone():
                stats["ar_seen"] += 1
                continue
            st = {"Open": "Open", "Paid": "Paid", "Doubtful": "Doubtful", "Estimated": "Estimate", "Estimate": "Estimate", "Partially Paid": "Partially Paid"}.get(status, "Open")
            job = None
            from .ledger.intake import _norm_job
            job = _norm_job(conn, None, f"{cust} {inv} {d.get('Notes') or ''}")
            db.insert(conn, "ar_invoices", dict(customer=cust, invoice_no=inv, invoice_date=_d(d.get("Invoice date")), due_date=_d(d.get("Expected collection date")),
                                                amount=amt, status=st, expected_date=_d(d.get("Expected collection date")), paid_date=_d(d.get("Date paid")),
                                                paid_amount=amt if st == "Paid" else 0, job_no=job, notes=f"seeded from cash flow tool. {str(d.get('Notes') or '')[:300]}", created_at=db.now_iso()))
            stats["ar_new"] += 1
    if "WEEKLY_INPUT" in wb.sheetnames:
        rows = list(wb["WEEKLY_INPUT"].iter_rows(values_only=True))
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        dated = [dict(zip(hdr, r)) for r in rows[1:] if r and _d(r[0])]
        if dated:
            last = dated[-1]
            as_of = _d(last["Date"])
            mapping = {"RBC chequing": "rbc_chq", "TD chequing": "td_chq", "TD Direct Investing": "td_di", "TD Bus Visa": "td_visa", "Op. line drawn (+)": "td_loc", "Capital One MC": "cap_one"}
            for col, key in mapping.items():
                v = _num(last.get(col))
                if v is None:
                    continue
                if key in ("td_visa", "cap_one"):
                    v = abs(v)   # store card balances as amount owing
                if db.upsert_ignore(conn, "bank_balances", dict(as_of=as_of, account_key=key, balance=v, source="manual", created_at=db.now_iso())):
                    stats["balances"] += 1
    conn.commit()
    return stats


# ---------------------------------------------------------------- config-driven
def seed_from_config(conn: sqlite3.Connection, settings) -> dict[str, int]:
    raw = settings.raw
    n = {"jobs": 0, "debts": 0, "recurring": 0, "employees": 0, "equipment": 0}
    for j in raw.get("jobs", []):
        conn.execute("INSERT INTO jobs(job_no,name,client,status,contract_value,sharepoint_folder) VALUES(?,?,?,?,?,?) ON CONFLICT(job_no) DO UPDATE SET name=excluded.name, client=excluded.client, status=excluded.status, contract_value=excluded.contract_value, sharepoint_folder=excluded.sharepoint_folder",
                     (str(j["job_no"]), j["name"], j.get("client"), j.get("status", "active"), j.get("contract_value"), j.get("sharepoint_folder")))
        n["jobs"] += 1
    for d in raw.get("debts", []):
        conn.execute("INSERT INTO debts(name,kind,balance,monthly_payment,payment_day,annual_rate,critical) VALUES(?,?,?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET kind=excluded.kind, monthly_payment=excluded.monthly_payment, payment_day=excluded.payment_day, annual_rate=excluded.annual_rate, critical=excluded.critical, balance=COALESCE(debts.balance, excluded.balance)",
                     (d["name"], d.get("kind"), d.get("balance"), d.get("monthly_payment", 0), d.get("payment_day"), d.get("annual_rate"), 1 if d.get("critical", True) else 0))
        n["debts"] += 1
    for r in raw.get("recurring", []):
        conn.execute("INSERT INTO recurring(name,amount,direction,cadence,next_date,category,critical,end_date) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(name) DO UPDATE SET amount=excluded.amount, direction=excluded.direction, cadence=excluded.cadence, category=excluded.category, critical=excluded.critical, end_date=excluded.end_date, next_date=CASE WHEN recurring.next_date < date('now') THEN excluded.next_date ELSE recurring.next_date END",
                     (r["name"], r["amount"], r["direction"], r["cadence"], str(r["next_date"]), r.get("category"), 1 if r.get("critical") else 0, str(r["end_date"]) if r.get("end_date") else None))
        n["recurring"] += 1
    for e in raw.get("employees", []):
        conn.execute("INSERT INTO employees(name,position,base_rate,active) VALUES(?,?,?,?) ON CONFLICT(name) DO UPDATE SET position=excluded.position, base_rate=COALESCE(excluded.base_rate, employees.base_rate), active=excluded.active",
                     (e["name"], e.get("position"), e.get("base_rate"), 1 if e.get("active", True) else 0))
        n["employees"] += 1
    for u in raw.get("equipment", []):
        conn.execute("INSERT INTO equipment(unit_id,description,make_model,sharepoint_folder) VALUES(?,?,?,?) ON CONFLICT(unit_id) DO UPDATE SET description=excluded.description, make_model=excluded.make_model, sharepoint_folder=COALESCE(excluded.sharepoint_folder, equipment.sharepoint_folder)",
                     (u["unit_id"], u.get("description"), u.get("make_model"), u.get("sharepoint_folder")))
        n["equipment"] += 1
    conn.commit()
    return n


def seed_from_sharepoint(conn: sqlite3.Connection, settings, graph, finance_drive: str, projects_drive: str | None) -> dict:
    out: dict = {}
    src = settings.raw.get("seed_sources", {})
    cc = src.get("cost_codes_xlsx")
    if cc and projects_drive:
        data = graph.download_path(projects_drive if cc.get("library", "projects") == "projects" else finance_drive, cc["path"])
        out["cost_codes"] = load_cost_codes_xlsx(conn, data) if data else "NOT FOUND"
    ap = src.get("ap_register_xlsx")
    if ap:
        data = graph.download_path(finance_drive, ap)
        out["ap_register"] = import_ap_register(conn, data) if data else "NOT FOUND"
    cf = src.get("cashflow_tool_xlsx")
    if cf:
        data = graph.download_path(finance_drive, cf)
        out["cashflow_tool"] = import_cashflow_tool(conn, data, settings) if data else "NOT FOUND"
    return out
