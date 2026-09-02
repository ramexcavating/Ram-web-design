"""Read Rodney's edits back out of the review workbook. Only the GREEN columns; only when a value changed."""
from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .. import db
from ..rules.filing import parse_date


def _d(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    p = parse_date(str(v))
    return p.isoformat() if p else None


def _yn(v) -> int | None:
    if v in (None, ""):
        return None
    return 1 if str(v).strip().lower() in ("yes", "y", "true", "1") else 0


def _num(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "").replace("$", ""))
    except ValueError:
        return None


def _rows(ws):
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in (None, ""):
            continue
        yield dict(zip(hdr, row))


def import_workbook(conn: sqlite3.Connection, path: str | Path) -> dict[str, int]:
    wb = load_workbook(path, data_only=True)
    stats = {"ap": 0, "ar": 0, "receipts": 0, "missing": 0, "timesheets": 0, "actions": 0}
    if "AP_TRACKER" in wb.sheetnames:
        for r in _rows(wb["AP_TRACKER"]):
            cur = conn.execute("SELECT * FROM ap_invoices WHERE id=?", (int(r["ID"]),)).fetchone()
            if not cur:
                continue
            upd = {}
            amt = _num(r.get("Amount"))
            if amt is not None and (cur["amount"] is None or abs(amt - cur["amount"]) > 0.005):
                upd["amount"] = amt
            conf = _yn(r.get("Amt Confirmed"))
            if conf is not None and conf != cur["amount_confirmed"]:
                upd["amount_confirmed"] = conf
            for col, key in (("Payment Status", "status"), ("Method", "method"), ("Job", "job_no"), ("Cost Code", "cost_code"), ("Notes", "notes")):
                v = r.get(col)
                v = str(v).strip() if v not in (None, "") else None
                if v is not None and v != cur[key]:
                    upd[key] = v
            for col, key in (("Planned Pay Date", "planned_pay_date"), ("Date Paid", "paid_date")):
                v = _d(r.get(col))
                if v is not None and v != cur[key]:
                    upd[key] = v
            if upd.get("status") == "Paid" and "paid_date" not in upd and not cur["paid_date"]:
                upd["paid_date"] = date.today().isoformat()
            if upd:
                sets = ", ".join(f"{k}=?" for k in upd)
                conn.execute(f"UPDATE ap_invoices SET {sets} WHERE id=?", (*upd.values(), cur["id"]))
                stats["ap"] += 1
    if "AR_TRACKER" in wb.sheetnames:
        for r in _rows(wb["AR_TRACKER"]):
            cur = conn.execute("SELECT * FROM ar_invoices WHERE id=?", (int(r["ID"]),)).fetchone()
            if not cur:
                continue
            upd = {}
            amt = _num(r.get("Amount"))
            if amt is not None and abs(amt - cur["amount"]) > 0.005:
                upd["amount"] = amt
            for col, key in (("Status", "status"), ("Job", "job_no"), ("Notes", "notes")):
                v = r.get(col)
                v = str(v).strip() if v not in (None, "") else None
                if v is not None and v != cur[key]:
                    upd[key] = v
            for col, key in (("Expected Date", "expected_date"), ("Date Paid", "paid_date")):
                v = _d(r.get(col))
                if v is not None and v != cur[key]:
                    upd[key] = v
            if upd.get("status") == "Paid":
                upd.setdefault("paid_date", date.today().isoformat())
                upd["paid_amount"] = upd.get("amount", cur["amount"])
            if upd:
                sets = ", ".join(f"{k}=?" for k in upd)
                conn.execute(f"UPDATE ar_invoices SET {sets} WHERE id=?", (*upd.values(), cur["id"]))
                stats["ar"] += 1
    if "RECEIPTS" in wb.sheetnames:
        for r in _rows(wb["RECEIPTS"]):
            cur = conn.execute("SELECT * FROM receipts WHERE id=?", (int(r["ID"]),)).fetchone()
            if not cur:
                continue
            upd = {}
            amt = _num(r.get("Amount"))
            if amt is not None and (cur["amount"] is None or abs(amt - cur["amount"]) > 0.005):
                upd["amount"] = amt
            for col, key in (("Job", "job_no"), ("Cost Code", "cost_code"), ("Equipment", "equipment_id")):
                v = r.get(col)
                v = str(v).strip() if v not in (None, "") else None
                if v is not None and v != cur[key]:
                    upd[key] = v
            for col, key in (("Reimbursable", "reimbursable"), ("Reimbursed", "reimbursed")):
                v = _yn(r.get(col))
                if v is not None and v != cur[key]:
                    upd[key] = v
            if upd:
                sets = ", ".join(f"{k}=?" for k in upd)
                conn.execute(f"UPDATE receipts SET {sets} WHERE id=?", (*upd.values(), cur["id"]))
                stats["receipts"] += 1
    if "MISSING_RECEIPTS" in wb.sheetnames:
        for r in _rows(wb["MISSING_RECEIPTS"]):
            if _yn(r.get("Personal / no receipt needed")):
                conn.execute("UPDATE bank_transactions SET receipt_required=0, match_type='rule', category=COALESCE(category,'personal') WHERE id=?", (int(r["Txn ID"]),))
                stats["missing"] += 1
    if "EQUIPMENT" in wb.sheetnames:
        for r in _rows(wb["EQUIPMENT"]):
            d = r.get("Description")
            if d not in (None, ""):
                conn.execute("INSERT INTO equipment(unit_id, description) VALUES(?,?) ON CONFLICT(unit_id) DO UPDATE SET description=excluded.description", (str(r["Unit"]).strip(), str(d).strip()))
    if "TIMESHEETS" in wb.sheetnames:
        for r in _rows(wb["TIMESHEETS"]):
            v = r.get("Status")
            if v not in (None, ""):
                cur = conn.execute("UPDATE timesheets SET status=? WHERE id=? AND status<>?", (str(v).strip(), int(r["ID"]), str(v).strip()))
                stats["timesheets"] += cur.rowcount
    if "ACTIONS" in wb.sheetnames:
        for r in _rows(wb["ACTIONS"]):
            st = str(r.get("Status") or "").strip().lower()
            ans = r.get("Your answer")
            if st in ("resolved", "dismissed"):
                conn.execute("UPDATE action_items SET status=?, resolved_at=?, detail=COALESCE(detail,'') || ? WHERE id=? AND status='open'",
                             (st, db.now_iso(), f"\nAnswer: {ans}" if ans else "", int(r["ID"])))
                stats["actions"] += 1
            elif ans not in (None, ""):
                conn.execute("UPDATE action_items SET detail=COALESCE(detail,'') || ? WHERE id=?", (f"\nAnswer: {ans}", int(r["ID"])))
                stats["actions"] += 1
    conn.commit()
    return stats
